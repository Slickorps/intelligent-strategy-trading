"""Standalone report generation module.

Generates HTML reports, PDF reports, Excel exports, and chart visualizations
for backtest results, portfolio snapshots, and risk assessments.

Usage:
    from ist.backtest.report import BacktestReporter

    reporter = BacktestReporter()

    # Generate HTML report
    html = reporter.generate_html_report(backtest_data, title="My Strategy")

    # Generate PDF report
    pdf_path = reporter.generate_pdf_report(backtest_data, output_path="report.pdf")

    # Export to Excel
    excel_path = reporter.export_to_excel(backtest_data, output_path="results.xlsx")

    # Generate equity curve chart
    chart_path = reporter.generate_equity_curve_chart(equity_curve, output_path="equity.png")
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from ist.core.logging import get_logger

logger = get_logger(__name__)


class BacktestReporter:
    """Generate backtest reports in multiple formats.

    Supports:
    - HTML: styled responsive report with metrics, tables, and charts
    - PDF: printable PDF report via fpdf2
    - Excel: .xlsx export with trades sheet and equity curve sheet
    - Charts: mplfinance OHLCV charts and equity curve plots

    Attributes:
        output_dir: Default output directory for generated files.
    """

    def __init__(self, output_dir: str = ".") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # HTML Report
    # ------------------------------------------------------------------

    def generate_html_report(
        self,
        data: dict[str, Any],
        title: str = "Backtest Report",
        report_type: str = "backtest",
    ) -> str:
        """Generate a styled HTML report.

        Args:
            data: Report data dictionary containing metrics, trades, equity_curve.
            title: Report title displayed at top of page.
            report_type: One of 'backtest', 'portfolio', 'risk'.

        Returns:
            Complete HTML document string.
        """
        template_map = {
            "backtest": self._backtest_html,
            "portfolio": self._portfolio_html,
            "risk": self._risk_html,
        }
        template = template_map.get(report_type, self._backtest_html)
        return template(data, title)

    # ------------------------------------------------------------------
    # PDF Report
    # ------------------------------------------------------------------

    def generate_pdf_report(
        self,
        data: dict[str, Any],
        output_path: str = "backtest_report.pdf",
        title: str = "Backtest Report",
    ) -> str:
        """Generate a PDF report using fpdf2.

        Args:
            data: Report data dictionary.
            output_path: File path for the generated PDF.
            title: Report title.

        Returns:
            Absolute path to the generated PDF file.
        """
        try:
            from fpdf import FPDF
        except ImportError:
            logger.error("fpdf2 is not installed. Run: pip install fpdf2")
            return ""

        full_path = self.output_dir / output_path

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # --- Helper to add a Unicode font, falling back to built-in ---
        _setup_pdf_font(pdf)

        # Title
        pdf.set_font("Helvetica", "B", 20)
        pdf.cell(0, 12, title, ln=True, align="C")
        pdf.ln(4)

        # Timestamp
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 8, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", ln=True)
        pdf.ln(4)
        pdf.set_text_color(0, 0, 0)

        metrics = data.get("metrics") or data.get("summary") or {}
        trades = data.get("trades", [])

        # ---- Section: Performance Summary ----
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Performance Summary", ln=True)
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 11)
        metric_rows = [
            ("Total Return", _fmt_pct(metrics.get("total_return", 0))),
            ("Annualized Return", _fmt_pct(metrics.get("annualized_return", 0))),
            ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}"),
            ("Sortino Ratio", f"{metrics.get('sortino_ratio', 0):.2f}"),
            ("Calmar Ratio", f"{metrics.get('calmar_ratio', 0):.2f}"),
            ("Max Drawdown", _fmt_pct(metrics.get("max_drawdown", 0))),
            ("Volatility (Ann.)", _fmt_pct(metrics.get("volatility", 0))),
            ("Win Rate", _fmt_pct(metrics.get("win_rate", 0))),
            ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}"),
        ]
        col_w = 95
        for label, value in metric_rows:
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(col_w, 8, label, border=0)
            pdf.set_font("Helvetica", "", 11)
            pdf.cell(0, 8, value, border=0, ln=True)

        pdf.ln(6)

        # ---- Section: Trade Statistics ----
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Trade Statistics", ln=True)
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 11)
        trade_rows = [
            ("Total Trades", str(metrics.get("total_trades", 0))),
            ("Winning Trades", str(metrics.get("winning_trades", 0))),
            ("Losing Trades", str(metrics.get("losing_trades", 0))),
            ("Avg Trade", f"${metrics.get('avg_trade', 0):,.2f}"),
            ("Avg Win", f"${metrics.get('avg_win', 0):,.2f}"),
            ("Avg Loss", f"${metrics.get('avg_loss', 0):,.2f}"),
        ]
        for label, value in trade_rows:
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(col_w, 8, label, border=0)
            pdf.set_font("Helvetica", "", 11)
            pdf.cell(0, 8, value, border=0, ln=True)

        pdf.ln(6)

        # ---- Section: Assessment ----
        assessment = data.get("assessment", "")
        if assessment:
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, "Assessment", ln=True)
            pdf.ln(2)
            pdf.set_font("Helvetica", "", 11)
            pdf.multi_cell(0, 7, assessment)
            pdf.ln(4)

        # ---- Section: Trade Log (first 50) ----
        if trades:
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, "Trade Log (first 50)", ln=True)
            pdf.ln(2)

            # Table header
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(74, 144, 217)
            pdf.set_text_color(255, 255, 255)
            headers = ["#", "Symbol", "Side", "Entry", "Exit", "PnL"]
            widths = [10, 28, 14, 28, 28, 28]
            for h, w in zip(headers, widths):
                pdf.cell(w, 7, h, border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_text_color(0, 0, 0)

            pdf.set_font("Helvetica", "", 8)
            for i, t in enumerate(trades[:50]):
                if i % 2 == 0:
                    pdf.set_fill_color(245, 245, 245)
                else:
                    pdf.set_fill_color(255, 255, 255)
                row_data = [
                    str(i + 1),
                    str(t.get("symbol", t.get("Symbol", "")))[:10],
                    str(t.get("side", t.get("Side", "")))[:5],
                    _fmt_price(t.get("entry_price", t.get("EntryPrice", ""))),
                    _fmt_price(t.get("exit_price", t.get("ExitPrice", ""))),
                    _fmt_pnl(t.get("realized_pnl", t.get("PnL", 0))),
                ]
                for d, w in zip(row_data, widths):
                    pdf.cell(w, 6, d, border=1, fill=True, align="C")
                pdf.ln()

        # Save
        pdf.output(str(full_path))
        logger.info(f"PDF report saved to {full_path}")
        return str(full_path.absolute())

    # ------------------------------------------------------------------
    # Excel Export
    # ------------------------------------------------------------------

    def export_to_excel(
        self,
        data: dict[str, Any],
        output_path: str = "backtest_results.xlsx",
    ) -> str:
        """Export backtest data to an Excel workbook.

        Creates sheets:
            - 'Summary': performance metrics as a two-column table.
            - 'Trades': trade log with all metadata.
            - 'Equity Curve': daily equity values and drawdown.

        Args:
            data: Report data dictionary.
            output_path: File path for the generated .xlsx file.

        Returns:
            Absolute path to the generated Excel file.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl is not installed. Run: pip install openpyxl")
            return ""

        full_path = self.output_dir / output_path
        wb = Workbook()

        # --- Styles ---
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_align = Alignment(horizontal="center")

        # ---- Sheet 1: Summary ----
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Metric", "Value"])

        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        metrics = data.get("metrics") or data.get("summary") or {}
        summary_rows = [
            ("Total Return", _fmt_pct(metrics.get("total_return", 0))),
            ("Annualized Return", _fmt_pct(metrics.get("annualized_return", 0))),
            ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}"),
            ("Sortino Ratio", f"{metrics.get('sortino_ratio', 0):.2f}"),
            ("Calmar Ratio", f"{metrics.get('calmar_ratio', 0):.2f}"),
            ("Max Drawdown", _fmt_pct(metrics.get("max_drawdown", 0))),
            ("Max DD Duration", f"{metrics.get('max_drawdown_duration', 0)} days"),
            ("Volatility (Ann.)", _fmt_pct(metrics.get("volatility", 0))),
            ("Win Rate", _fmt_pct(metrics.get("win_rate", 0))),
            ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}"),
            ("Recovery Factor", f"{metrics.get('recovery_factor', 0):.2f}"),
            ("Total Trades", str(metrics.get("total_trades", 0))),
            ("Winning Trades", str(metrics.get("winning_trades", 0))),
            ("Losing Trades", str(metrics.get("losing_trades", 0))),
            ("Avg Trade", f"${metrics.get('avg_trade', 0):,.2f}"),
            ("Avg Win", f"${metrics.get('avg_win', 0):,.2f}"),
            ("Avg Loss", f"${metrics.get('avg_loss', 0):,.2f}"),
        ]
        for row in summary_rows:
            ws_summary.append(row)

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20

        # ---- Sheet 2: Trades ----
        trades = data.get("trades", [])
        if trades:
            ws_trades = wb.create_sheet("Trades")
            if trades:
                trade_headers = list(trades[0].keys())
            else:
                trade_headers = ["symbol", "side", "entry_price", "exit_price", "realized_pnl"]
            ws_trades.append(trade_headers)

            for cell in ws_trades[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            for t in trades:
                row = [t.get(h, "") for h in trade_headers]
                ws_trades.append(row)

            # Auto-fit column widths
            for col_idx, _ in enumerate(trade_headers, start=1):
                ws_trades.column_dimensions[
                    ws_trades.cell(row=1, column=col_idx).column_letter
                ].width = 18

        # ---- Sheet 3: Equity Curve ----
        equity_curve = data.get("equity_curve", [])
        if equity_curve:
            ws_equity = wb.create_sheet("Equity Curve")
            ws_equity.append(["Date", "Equity", "Daily Return", "Drawdown"])

            for cell in ws_equity[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            peak = 0
            for entry in equity_curve:
                date = entry.get("date", "")
                equity = entry.get("equity", 0)
                if equity > peak:
                    peak = equity
                dd = (peak - equity) / peak if peak > 0 else 0
                ws_equity.append([date, equity, None, round(dd, 6)])

            ws_equity.column_dimensions["A"].width = 20
            ws_equity.column_dimensions["B"].width = 16
            ws_equity.column_dimensions["C"].width = 16
            ws_equity.column_dimensions["D"].width = 14

        wb.save(str(full_path))
        logger.info(f"Excel report saved to {full_path}")
        return str(full_path.absolute())

    # ------------------------------------------------------------------
    # Chart Generation
    # ------------------------------------------------------------------

    def generate_equity_curve_chart(
        self,
        equity_curve: list[dict],
        output_path: str = "equity_curve.png",
        title: str = "Equity Curve",
        figsize: tuple = (14, 6),
    ) -> str:
        """Generate an equity curve and drawdown chart using matplotlib.

        Produces a two-panel figure:
            Top panel: equity curve with peak overlay.
            Bottom panel: drawdown percentage.

        Args:
            equity_curve: List of dicts with 'date' and 'equity' keys.
            output_path: File path for the generated PNG.
            title: Chart title.
            figsize: Figure size as (width, height) in inches.

        Returns:
            Absolute path to the generated chart PNG file.
        """
        try:
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
        except ImportError:
            logger.error("matplotlib is not installed. Run: pip install matplotlib")
            return ""

        if not equity_curve:
            logger.warning("Empty equity curve, skipping chart generation.")
            return ""

        full_path = self.output_dir / output_path

        dates = []
        equity_values = []

        for entry in equity_curve:
            d = entry.get("date", "")
            if isinstance(d, str):
                try:
                    d = pd.Timestamp(d)
                except Exception:
                    d = datetime.fromisoformat(d)
            dates.append(d)
            equity_values.append(entry.get("equity", 0))

        # Calculate drawdown
        peak = equity_values[0]
        drawdowns = []
        for v in equity_values:
            if v > peak:
                peak = v
            dd = (peak - v) / peak if peak > 0 else 0
            drawdowns.append(-dd * 100)

        fig, (ax1, ax2) = plt.subplots(
            2, 1, figsize=figsize, sharex=True,
            gridspec_kw={"height_ratios": [3, 1]}
        )

        # Equity curve
        ax1.plot(dates, equity_values, color="#4A90D9", linewidth=1.5, label="Equity")
        ax1.plot(dates, [peak] * len(dates), color="gray", linestyle="--", alpha=0.5, label="Peak")
        ax1.fill_between(dates, equity_values, alpha=0.1, color="#4A90D9")
        ax1.set_title(title, fontsize=14, fontweight="bold")
        ax1.set_ylabel("Equity ($)")
        ax1.legend(loc="upper left")
        ax1.grid(True, alpha=0.3)

        # Drawdown
        ax2.fill_between(dates, drawdowns, 0, color="#E74C3C", alpha=0.6)
        ax2.set_ylabel("Drawdown (%)")
        ax2.set_xlabel("Date")
        ax2.grid(True, alpha=0.3)

        # Date formatting
        ax2.xaxis.set_major_formatter(mdates.AutoDateFormatter(mdates.AutoDateLocator()))
        fig.autofmt_xdate()

        plt.tight_layout()
        plt.savefig(str(full_path), dpi=150, bbox_inches="tight")
        plt.close(fig)
        logger.info(f"Equity curve chart saved to {full_path}")
        return str(full_path.absolute())

    def generate_ohlcv_chart(
        self,
        ohlcv_data: pd.DataFrame,
        output_path: str = "ohlcv_chart.png",
        title: str = "OHLCV Chart",
        volume: bool = True,
    ) -> str:
        """Generate a candlestick chart with optional volume using mplfinance.

        Args:
            ohlcv_data: DataFrame with columns Open, High, Low, Close, Volume
                        and a DatetimeIndex.
            output_path: File path for the generated PNG.
            title: Chart title.
            volume: Whether to include volume subplot.

        Returns:
            Absolute path to the generated chart PNG file.
        """
        try:
            import mplfinance as mpf
        except ImportError:
            logger.error("mplfinance is not installed. Run: pip install mplfinance")
            return ""

        if ohlcv_data.empty:
            logger.warning("Empty OHLCV data, skipping chart generation.")
            return ""

        full_path = self.output_dir / output_path

        # Ensure the DataFrame has the expected column names
        required_cols = {"Open", "High", "Low", "Close", "Volume"}
        if not required_cols.issubset(set(ohlcv_data.columns)):
            # Try to map lowercase columns
            col_map = {
                "open": "Open", "high": "High", "low": "Low",
                "close": "Close", "volume": "Volume",
            }
            renamed = {}
            for k, v in col_map.items():
                if k in ohlcv_data.columns and v not in ohlcv_data.columns:
                    renamed[k] = v
            if renamed:
                ohlcv_data = ohlcv_data.rename(columns=renamed)

        style = mpf.make_mpf_style(
            base_mpf_style="charles",
            gridcolor="#E0E0E0",
            facecolor="#FFFFFF",
        )

        kwargs: dict[str, Any] = {
            "type": "candle",
            "style": style,
            "title": title,
            "ylabel": "Price",
            "savefig": str(full_path),
        }
        if volume and "Volume" in ohlcv_data.columns:
            kwargs["volume"] = True

        mpf.plot(ohlcv_data, **kwargs)
        logger.info(f"OHLCV chart saved to {full_path}")
        return str(full_path.absolute())

    # ------------------------------------------------------------------
    # JSON / CSV Export
    # ------------------------------------------------------------------

    def export_to_json(self, data: dict[str, Any], indent: int = 2) -> str:
        """Export data as a JSON string."""
        return json.dumps(data, indent=indent, default=str)

    def export_trades_to_csv(self, trades: list[dict]) -> str:
        """Export trades to CSV string."""
        if not trades:
            return ""
        headers = list(trades[0].keys())
        lines = [",".join(headers)]
        for trade in trades:
            row = []
            for key in headers:
                value = trade.get(key, "")
                if isinstance(value, str) and ("," in value or '"' in value):
                    value = '"' + value.replace('"', '""') + '"'
                row.append(str(value))
            lines.append(",".join(row))
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Private HTML Templates
    # ------------------------------------------------------------------

    @staticmethod
    def _backtest_html(data: dict[str, Any], title: str) -> str:
        """HTML template for backtest report."""
        metrics = data.get("metrics") or data.get("summary") or {}
        trades = data.get("trades", [])

        # Build metric cards
        metric_cards = _render_metric_cards_html(metrics)
        trade_table = _render_trade_table_html(metrics)
        assessment = data.get("assessment", "")

        return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        :root {{
            --primary: #4a90d9;
            --bg: #f5f7fa;
            --card-bg: #ffffff;
            --text: #333333;
            --text-muted: #666666;
            --positive: #27ae60;
            --negative: #e74c3c;
            --border: #e0e0e0;
        }}
        body {{
            font-family: 'Segoe UI', Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background: var(--bg);
            color: var(--text);
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        h1 {{
            color: var(--primary);
            border-bottom: 3px solid var(--primary);
            padding-bottom: 12px;
            margin-bottom: 6px;
        }}
        .timestamp {{
            color: var(--text-muted);
            font-size: 13px;
            margin-bottom: 24px;
        }}
        h2 {{
            color: #444;
            margin-top: 32px;
            border-bottom: 1px solid var(--border);
            padding-bottom: 8px;
        }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
            gap: 16px;
            margin: 20px 0;
        }}
        .metric-card {{
            background: var(--card-bg);
            padding: 18px;
            border-radius: 10px;
            box-shadow: 0 1px 4px rgba(0,0,0,0.08);
            border-left: 5px solid var(--primary);
        }}
        .metric-label {{
            font-size: 12px;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .metric-value {{
            font-size: 22px;
            font-weight: 700;
            color: var(--text);
            margin-top: 6px;
        }}
        .metric-value.positive {{ color: var(--positive); }}
        .metric-value.negative {{ color: var(--negative); }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background: var(--card-bg);
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 1px 4px rgba(0,0,0,0.06);
        }}
        th, td {{
            padding: 12px 16px;
            text-align: left;
        }}
        th {{
            background: var(--primary);
            color: white;
            font-weight: 600;
            font-size: 14px;
        }}
        tr:nth-child(even) td {{
            background: #f8f9fa;
        }}
        .assessment {{
            background: #eef5fc;
            border-left: 4px solid var(--primary);
            padding: 14px 18px;
            margin: 16px 0;
            border-radius: 4px;
            line-height: 1.6;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 16px;
            border-top: 1px solid var(--border);
            color: var(--text-muted);
            font-size: 12px;
            text-align: center;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <p class="timestamp">Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</p>

        <h2>Performance Summary</h2>
        <div class="metrics-grid">
            {metric_cards}
        </div>

        <h2>Trade Statistics</h2>
        {trade_table}

        {_assessment_html(assessment)}

        <div class="footer">
            Generated by Intelligent Strategy Trading Platform &mdash; Backtest Reporter
        </div>
    </div>
</body>
</html>"""

    @staticmethod
    def _portfolio_html(data: dict[str, Any], title: str) -> str:
        """HTML template for portfolio report."""
        return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; background: #f5f7fa; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        h1 {{ color: #4a90d9; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <p><strong>Portfolio Value:</strong> ${data.get('equity', 0):,.2f}</p>
        <p><strong>Cash:</strong> ${data.get('cash', 0):,.2f}</p>
        <p><strong>Open Positions:</strong> {len(data.get('positions', []))}</p>
    </div>
</body>
</html>"""

    @staticmethod
    def _risk_html(data: dict[str, Any], title: str) -> str:
        """HTML template for risk report."""
        var_95 = data.get("var_95", 0)
        cvar_95 = data.get("cvar_95", 0)
        return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; background: #f5f7fa; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        h1 {{ color: #4a90d9; }}
        .risk-metric {{ background: #f8f9fa; padding: 16px; margin: 10px 0; border-radius: 6px; border-left: 4px solid #e74c3c; }}
        .risk-metric h3 {{ margin-top: 0; color: #e74c3c; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <div class="risk-metric">
            <h3>Value at Risk (95%)</h3>
            <p>{var_95:.2%}</p>
        </div>
        <div class="risk-metric">
            <h3>Conditional VaR (95%)</h3>
            <p>{cvar_95:.2%}</p>
        </div>
    </div>
</body>
</html>"""


# ------------------------------------------------------------------
# HTML helper functions
# ------------------------------------------------------------------

def _render_metric_cards_html(metrics: dict[str, Any]) -> str:
    """Build metric card HTML snippets."""
    cards = [
        ("Total Return", metrics.get("total_return", 0), True),
        ("Annualized Return", metrics.get("annualized_return", 0), True),
        ("Sharpe Ratio", metrics.get("sharpe_ratio", 0), False),
        ("Sortino Ratio", metrics.get("sortino_ratio", 0), False),
        ("Max Drawdown", metrics.get("max_drawdown", 0), True),
        ("Win Rate", metrics.get("win_rate", 0), True),
        ("Profit Factor", metrics.get("profit_factor", 0), False),
        ("Total Trades", metrics.get("total_trades", 0), False),
    ]

    html_parts = []
    for label, value, is_pct in cards:
        if is_pct and isinstance(value, (int, float)):
            cls = "positive" if value > 0 else "negative"
            display = f"{value:.2%}"
        elif isinstance(value, float):
            cls = ""
            display = f"{value:.2f}"
        else:
            cls = ""
            display = str(value)
        html_parts.append(
            f"""<div class="metric-card">
                <div class="metric-label">{label}</div>
                <div class="metric-value {cls}">{display}</div>
            </div>"""
        )
    return "\n".join(html_parts)


def _render_trade_table_html(metrics: dict[str, Any]) -> str:
    """Build trade statistics HTML table."""
    rows_html = []
    trade_metrics = [
        ("Total Trades", metrics.get("total_trades", 0)),
        ("Winning Trades", metrics.get("winning_trades", 0)),
        ("Losing Trades", metrics.get("losing_trades", 0)),
        ("Win Rate", f"{metrics.get('win_rate', 0):.1%}"),
        ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}"),
        ("Avg Trade", f"${metrics.get('avg_trade', 0):,.2f}"),
        ("Avg Win", f"${metrics.get('avg_win', 0):,.2f}"),
        ("Avg Loss", f"${metrics.get('avg_loss', 0):,.2f}"),
        ("Expectancy", f"${metrics.get('expectancy', 0):,.2f}"),
    ]
    for label, value in trade_metrics:
        rows_html.append(f"<tr><td>{label}</td><td>{value}</td></tr>")

    return f"""<table>
        <thead><tr><th>Metric</th><th>Value</th></tr></thead>
        <tbody>{"".join(rows_html)}</tbody>
    </table>"""


def _assessment_html(assessment: str) -> str:
    """Build assessment HTML block."""
    if not assessment:
        return ""
    return f'<div class="assessment"><strong>Assessment:</strong> {assessment}</div>'


# ------------------------------------------------------------------
# PDF helper
# ------------------------------------------------------------------

def _setup_pdf_font(pdf: Any) -> None:
    """Configure a Unicode-capable font for the PDF.

    Attempts to use DejaVu (bundled with fpdf2) for CJK support,
    falling back to Helvetica.
    """
    try:
        pdf.add_font("DejaVu", "", r"C:\Windows\Fonts\DejaVuSans.ttf", uni=True)
        pdf.add_font("DejaVu", "B", r"C:\Windows\Fonts\DejaVuSans-Bold.ttf", uni=True)
        pdf.set_font("DejaVu", "", 11)
    except Exception:
        pass


# ------------------------------------------------------------------
# Formatting helpers
# ------------------------------------------------------------------

def _fmt_pct(value: Any) -> str:
    """Safely format a value as percentage string."""
    try:
        v = float(value)
        return f"{v:.2%}"
    except (TypeError, ValueError):
        return str(value)


def _fmt_price(value: Any) -> str:
    """Safely format a price value."""
    try:
        v = float(value)
        return f"${v:,.4f}"
    except (TypeError, ValueError):
        return str(value)


def _fmt_pnl(value: Any) -> str:
    """Safely format a PnL value."""
    try:
        v = float(value)
        return f"${v:,.2f}"
    except (TypeError, ValueError):
        return str(value)