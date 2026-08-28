"""Unit tests for the backtest report generator."""

import json
import sys
from pathlib import Path

import pandas as pd
import pytest

from ist.backtest.report import (
    BacktestReporter,
    _fmt_pct,
    _fmt_pnl,
    _fmt_price,
)


@pytest.fixture
def reporter(tmp_path) -> BacktestReporter:
    return BacktestReporter(output_dir=str(tmp_path))


@pytest.fixture
def sample_data() -> dict:
    return {
        "metrics": {
            "total_return": 0.15,
            "sharpe_ratio": 1.5,
            "max_drawdown": -0.05,
            "win_rate": 0.6,
            "profit_factor": 1.8,
            "total_trades": 10,
        },
        "trades": [
            {
                "symbol": "EURUSD",
                "side": "buy",
                "entry_price": 1.0,
                "exit_price": 1.1,
                "realized_pnl": 100.0,
            },
            {
                "symbol": "GBPUSD",
                "side": "sell",
                "entry_price": 1.3,
                "exit_price": 1.28,
                "realized_pnl": -20.0,
            },
        ],
        "equity_curve": [
            {"date": "2024-01-01", "equity": 100000.0},
            {"date": "2024-01-02", "equity": 101000.0},
            {"date": "2024-01-03", "equity": 100500.0},
        ],
        "assessment": "Strategy performed well.",
    }


class TestHtmlReport:
    def test_backtest_html(self, reporter, sample_data) -> None:
        html = reporter.generate_html_report(sample_data, title="My Strategy")
        assert html.startswith("<!DOCTYPE html>")
        assert "My Strategy" in html
        assert "Total Return" in html
        assert "Sharpe Ratio" in html

    def test_default_title(self, reporter) -> None:
        html = reporter.generate_html_report({"metrics": {}})
        assert "Backtest Report" in html

    def test_portfolio_html(self, reporter) -> None:
        html = reporter.generate_html_report(
            {"equity": 100000, "cash": 50000, "positions": []},
            report_type="portfolio",
        )
        assert "Portfolio Value" in html

    def test_risk_html(self, reporter) -> None:
        html = reporter.generate_html_report(
            {"var_95": 0.05, "cvar_95": 0.08},
            report_type="risk",
        )
        assert "Value at Risk" in html

    def test_unknown_type_falls_back_to_backtest(self, reporter, sample_data) -> None:
        html = reporter.generate_html_report(sample_data, report_type="unknown")
        assert "Performance Summary" in html

    def test_includes_assessment(self, reporter, sample_data) -> None:
        html = reporter.generate_html_report(sample_data)
        assert "Strategy performed well." in html


class TestJsonExport:
    def test_export_to_json(self, reporter, sample_data) -> None:
        parsed = json.loads(reporter.export_to_json(sample_data))
        assert parsed["metrics"]["sharpe_ratio"] == 1.5

    def test_export_to_json_indent(self, reporter) -> None:
        result = reporter.export_to_json({"a": 1}, indent=4)
        assert "\n    " in result


class TestCsvExport:
    def test_export_trades_to_csv(self, reporter, sample_data) -> None:
        csv = reporter.export_trades_to_csv(sample_data["trades"])
        lines = csv.splitlines()
        assert len(lines) == 3
        assert lines[0] == "symbol,side,entry_price,exit_price,realized_pnl"

    def test_export_trades_to_csv_empty(self, reporter) -> None:
        assert reporter.export_trades_to_csv([]) == ""

    def test_export_trades_to_csv_quotes_commas(self, reporter) -> None:
        csv = reporter.export_trades_to_csv([{"note": "a,b", "value": 1}])
        assert '"a,b"' in csv


class TestPdfReport:
    def test_generate_pdf_report(self, reporter, sample_data) -> None:
        path = reporter.generate_pdf_report(
            sample_data, output_path="report.pdf", title="My Report",
        )
        p = Path(path)
        assert p.exists()
        assert p.read_bytes().startswith(b"%PDF")

    def test_generate_pdf_report_empty_metrics(self, reporter) -> None:
        path = reporter.generate_pdf_report({}, output_path="empty.pdf")
        assert Path(path).exists()

    def test_generate_pdf_report_missing_fpdf(self, reporter, monkeypatch) -> None:
        monkeypatch.setitem(sys.modules, "fpdf", None)
        assert reporter.generate_pdf_report({"metrics": {}}) == ""


class TestExcelExport:
    def test_export_to_excel(self, reporter, sample_data) -> None:
        path = reporter.export_to_excel(sample_data, output_path="results.xlsx")

        from openpyxl import load_workbook

        wb = load_workbook(path)
        assert "Summary" in wb.sheetnames
        assert "Trades" in wb.sheetnames
        assert "Equity Curve" in wb.sheetnames

    def test_export_to_excel_missing_openpyxl(self, reporter, monkeypatch) -> None:
        monkeypatch.setitem(sys.modules, "openpyxl", None)
        assert reporter.export_to_excel({"metrics": {}}) == ""


class TestCharts:
    def test_equity_curve_chart_empty_returns_empty(self, reporter) -> None:
        assert reporter.generate_equity_curve_chart([]) == ""

    def test_equity_curve_chart(self, reporter, sample_data, monkeypatch) -> None:
        monkeypatch.setenv("MPLBACKEND", "Agg")
        path = reporter.generate_equity_curve_chart(
            sample_data["equity_curve"], output_path="equity.png",
        )
        p = Path(path)
        assert p.exists()
        assert p.read_bytes().startswith(b"\x89PNG")

    def test_ohlcv_chart_empty_returns_empty(self, reporter) -> None:
        assert reporter.generate_ohlcv_chart(pd.DataFrame()) == ""

    def test_ohlcv_chart(self, reporter, monkeypatch) -> None:
        monkeypatch.setenv("MPLBACKEND", "Agg")
        idx = pd.date_range("2024-01-01", periods=5, freq="D")
        df = pd.DataFrame(
            {
                "Open": [100.0, 101.0, 102.0, 103.0, 104.0],
                "High": [105.0, 106.0, 107.0, 108.0, 109.0],
                "Low": [99.0, 100.0, 101.0, 102.0, 103.0],
                "Close": [104.0, 105.0, 106.0, 107.0, 108.0],
                "Volume": [1000, 1100, 1200, 1300, 1400],
            },
            index=idx,
        )
        path = reporter.generate_ohlcv_chart(df, output_path="ohlcv.png")
        assert Path(path).exists()


class TestFormatHelpers:
    def test_fmt_pct(self) -> None:
        assert _fmt_pct(0.1234) == "12.34%"
        assert _fmt_pct("not a number") == "not a number"

    def test_fmt_price(self) -> None:
        assert _fmt_price(1.2345) == "$1.2345"
        assert _fmt_price("") == ""

    def test_fmt_pnl(self) -> None:
        assert _fmt_pnl(123.456) == "$123.46"
        assert _fmt_pnl(None) == "None"
